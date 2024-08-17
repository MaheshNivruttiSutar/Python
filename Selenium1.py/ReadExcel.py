#Data Driven Testing Using Excel:
#                     - Reading data from Excel File
#                     - Writing Data into Excel File
#                     - Data Driven Test Cases

import openpyxl

#path="C:\Users\Mahesh Sutar\PycharmProjects\Selenium1.py\ReadExcel.xlsx"
path="C:\Drivers\ReadExcel.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active   #workbook.get_sheet_by_name("Sheet")

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r,column=c).value,end="        ")

    print()